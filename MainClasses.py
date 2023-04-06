
from openpyxl.reader.excel import load_workbook
from pythonds import Stack
import openpyxl as xl
import pydot

class ADF:
    """
    A class used to represent the ADF graph

    Attributes
    ----------
    name : str
        the name of the ADF
    nodes : dict
        the nodes which constitute the ADF
    reject : bool, default False
        is set to true when the reject keyword is used which lets the software know to reject the node rather than accep it when the condition is true
    nonLeaf : dict
        the nodes which are non-leaf that have children
    questionOrder : list
        an ordered list which determines which order the questions are asked in
    question : str, optional
        if the node is a base-level factor this stores the question
    statements : list
        the statements to be shown if the node is accepted or rejected
    nodeDone : list
        nodes which have been evaluated
    case : list
        the list of factors forming the case    
    
    Methods
    -------
    addNodes(name, acceptance = None, statement=None, question=None)
        allows nodes to be added to the ADF from the Node() class
    addMulti(name, acceptance, statement, question)
        allows nodes to be added to the ADF from the MultiChoice() class
    nonLeafGen()
        determines what is a non-leaf factor
    evaluateTree(case)
        evaluates the ADF for a specified case
    evaluateNode(node)
        evaluates the acceptance conditions of the node
    postfixEvaluation(acceptance)
        evaluates the individual acceptance conditions which are in postfix notation
    checkCondition(operator, op1, op2 = None):
        checks the logical conditions for the acceptance condition, returning a boolean
    checkNonLeaf(node)
        checks if a node has children which need to be evaluated before it is evaluated
    questionAssignment()
        checks if any node requires a question to be assigned
    visualiseNetwork(case=None)
        allows visualisation of the ADF
    saveNew(name)
        allows the ADF to be saved as a .xlsx file
    saveHelper(wb,name)
        helper class for saveNew which provides core functionality
    """
    
    def __init__(self, name):
        """
        Parameters
        ----------
        name : str
            the name of the ADF
        """
      
        self.name = name
        
        #dictionary of nodes --> 'name': 'node object
        self.nodes = {}
        
        self.reject = False
        
        #dictionary of nodes which have children
        self.nonLeaf = {}
        
        self.questionOrder = []
        
    def addNodes(self, name, acceptance = None, statement=None, question=None):
        """
        adds nodes to ADF
        
        Parameters
        ----------
        name : str
            the name of the node
        acceptance : list
            a list of the acceptance conditions each of which should be a string
        statement : list
            a list of the statements which will be shown if a condition is accepted or rejected
        question : str
            the question to determine whether a node is absent or present
        
        """
        
        node = Node(name, acceptance, statement, question)
        
        self.nodes[name] = node
        
        self.question = question
        
        #creates children nodes
        if node.children != None:
            for childName in node.children:
                if childName not in self.nodes:
                    node = Node(childName)
                    self.nodes[childName] = node
    
    def addMulti(self,name, acceptance, statement, question):
        """
        adds MultiChoice() nodes to the ADF
        
        Parameters
        ----------
        name : str
            the name of the node
        acceptance : list
            a list of the acceptance conditions each of which should be a string
        statement : list
            a list of the statements which will be shown if a condition is accepted or rejected
        question : str
            the question to determine whether a node is absent or present
        
        """
    
        node = MultiChoice(name, acceptance, statement, question)
        
        self.nodes[name] = node        
    
    def nonLeafGen(self):
        """
        determines which of the nodes is non-leaf        
        """
        
        #sets it back to an empty dictionary
        self.nonLeaf = {}
        
        #checks each node and determines if it is a non-leaf node (one with children)
        for name,node in zip(self.nodes,self.nodes.values()):
            
            #adds node to dict of nodes with children
            if node.children != None and node.children != []:
                self.nonLeaf[name] = node
            else:
                pass
                   
    def evaluateTree(self, case):
        """
        evaluates the ADF for a given case
        
        Parameters
        ----------
        case : list
            the list of factors forming the case 
        
        """
        #keep track of print statements
        self.statements = []
        #list of non-leaf nodes which have been evaluated
        self.nodeDone = []
        self.case = case
        #generates the non-leaf nodes
        self.nonLeafGen()
        #while there are nonLeaf nodes which have not been evaluated, evaluate a node in this list in ascending order  
        while self.nonLeaf != {}:
            for name,node in zip(self.nonLeaf,self.nonLeaf.values()):
                #checks if the node's children are non-leaf nodes
                if name == 'Decide' and len(self.nonLeaf) != 1:
                    pass     
                elif self.checkNonLeaf(node):
                    #adds to list of evaluated nodes
                    self.nodeDone.append(name) 
                    #checks candidate node's acceptance conditions
                    if self.evaluateNode(node):
                        #enables rejection clauses - handy for automobile
                        if self.reject != True:
                            #adds factor to case if present
                            self.case.append(name)
                        #deletes node from nonLeaf nodes
                        self.nonLeaf.pop(name)
                        self.statements.append(node.statement[self.counter])
                        self.reject = False
                        break
                    #if node's acceptance conditions are false                       
                    else:
                        #deletes node from nonLeaf nodes but doesn't add to case
                        self.nonLeaf.pop(name)
                        #the last statement is always the rejection statemenr
                        self.statements.append(node.statement[-1])
                        self.reject = False
                        break
        
        return self.statements
                                  
    def evaluateNode(self, node):
        """
        evaluates a node in respect to its acceptance conditions
        
        x will be always be a boolean value
        
        Parameters
        ----------
        node : class
            the node class to be evaluated
        
        """
        
        #for visualisation purposes - this tracks the attacking nodes
        self.vis = []
        
        #counter to index the statements to be shown to the user
        self.counter = -1
        
        #checks each acceptance condition seperately
        for i in node.acceptance:
            self.reject = False
            self.counter+=1
            x = self.postfixEvaluation(i)
            if x == True:
                return x  

        return x
    
    def postfixEvaluation(self,acceptance):
        """
        evaluates the given acceptance condition 
        
        Parameters
        ----------
        acceptance : str
            a string with the names of nodes seperated by logical operators            
        
        """
        #initialises stack of operands
        operandStack = Stack()
        #list of tokens from acceptance conditions
        tokenList = acceptance.split()
        #checks each token's acceptance conditions
        for token in tokenList:
            #checks if something is a rejection condition
            if token == 'reject':
                self.reject = True
                try:
                    if x in self.case:
                        return True
                    else:
                        return False
                except:
                    pass          
            elif token == 'not':
                operand1 = operandStack.pop()
                result = self.checkCondition(token,operand1)
                operandStack.push(result)
                self.vis.append(operand1)
                
            elif token == 'and' or token == 'or':
                operand2 = operandStack.pop()
                operand1 = operandStack.pop()
                result = self.checkCondition(token,operand1,operand2)
                operandStack.push(result)    
                                
            #for an acceptance condition with no operator 
            elif len(tokenList) == 1 or (len(tokenList) ==2 and 'reject' in tokenList):
                if 'reject' in tokenList and 'reject' != token:
                    x = token   
                else:
                    if token in self.case:
                        return True
                    else:
                        return False
            else:
                #adds the operand to the stack
                operandStack.push(token)
        
        return operandStack.pop()

    def checkCondition(self, operator, op1, op2 = None):
        """
        checks the logical condition and returns a boolean
        
        Parameters
        ----------
        operator : str
            the logical operator such as or, and, not  
        op1 : str
            the first operand
        op2 : str, optional
            the second operand
        """
        
        if operator == "or":
            if op1 in self.case or op2 in self.case or op1 == True or op2 == True:
                return True
            else:
                return False
            
        elif operator == "and":
            if op1 == True or op1 in self.case:
                if op2 in self.case or op2 == True:
                    return True 
                else: 
                    return False
            elif op2 == True or op2 in self.case:
                if op1 in self.case or op1 == True:
                    return True
                else:
                    return False   
            else:
                return False
            
        elif operator == "not":
            if op1 == True:
                return False
            if op1 == False:
                return True
            elif op1 not in self.case:
                return True
            else:
                return False
        
    def checkNonLeaf(self, node):
        """
        checks if a given node has children which need to be evaluated 
        before it can be evaluated
        
        Parameters
        ----------
        node : class
            the node class to be evaluated
        
        """
        for j in node.children:
    
            if j in self.nonLeaf:
                
                if j in self.nodeDone:
                    pass
                
                else:
                    return False

            else:
                pass

        return True
    
    def questionAssignment(self):
        """
        used by the user interface to determine whether a node needs a
        question assigning to it
        """
        for i in self.nodes.values():
            
            if i.children == None and i.question == None:
                
                return i.name  
            
        return None

    def visualiseNetwork(self,case=None):    
        """
        allows the ADF to be visualised as a graph
        
        can be for the domain with or without a case
        
        if there is a case it will highlight the nodes green which have been
        accepted and red the ones which have been rejected        
        
        Parameters
        ----------
        case : list, optional
            the list of factors constituting the case
        """
        
        #initialises the graph
        G = pydot.Dot('{}'.format(self.name), graph_type='graph')

        if case != None:

            #checks each node
            for i in self.nodes.values():
                
                #checks if node is already in the graph
                if i not in G.get_node_list():
                    
                    #checks if the node was accepted in the case
                    if i.name in case:
                        a = pydot.Node(i.name,label=i.name,color='green')
                    else:
                        a = pydot.Node(i.name,label=i.name,color='red')
                                        
                    G.add_node(a)
                
                #creates edges between a node and its children
                if i.children != None and i.children != []:
                    
                    self.evaluateNode(i)

                    for j in i.children:
                        
                                                
                        if j not in G.get_node_list():
                            
                            if j in case:
                                a = pydot.Node(j,label=j,color='green')
                            else:
                                a = pydot.Node(j,label=j,color='red')
                            
                            G.add_node(a)
                        
                        #self.vis is a list which tracks whether a node is an attacking or defending node
                        if j in self.vis:
                            if j in case:
                                my_edge = pydot.Edge(i.name, j, color='green',label='-')
                            else:
                                my_edge = pydot.Edge(i.name, j, color='red',label='-')
                        else:
                            if j in case:
                                my_edge = pydot.Edge(i.name, j, color='green',label='+')
                            else:
                                my_edge = pydot.Edge(i.name, j, color='red',label='+')

                        G.add_edge(my_edge)
        
        else:
            
            #creates self.vis if not already created
            self.evaluateTree([])
            
            #checks each node
            for i in self.nodes.values():
                
                #checks if node is already in the graph
                if i not in G.get_node_list():
                    
                    a = pydot.Node(i.name,label=i.name,color='black')

                    G.add_node(a)
                
                #creates edges between a node and its children
                if i.children != None and i.children != []:
                    
                    self.evaluateNode(i)

                    for j in i.children:
                        
                        if j not in G.get_node_list():
                            
                            a = pydot.Node(j,label=j,color='black')
                           
                            G.add_node(a)
                        
                        #self.vis is a list which tracks whether a node is an attacking or defending node
                        if j in self.vis:
                            my_edge = pydot.Edge(i.name, j, color='black',label='-')

                        else:
                            my_edge = pydot.Edge(i.name, j, color='black',label='+')

                        G.add_edge(my_edge)
        
        return G
    
    def saveNew(self,filename):
        """
        enables an ADF to be saved as a .xlsx file 
        
        Parameters
        ----------
        adf : class
            the instance of the ADF() class
        name : str
            the filename
        """
        
        #create the excel workbook
        wb = xl.Workbook()
        
        #saves the workbook
        wb.save('{}.xlsx'.format(filename))
        
        #saves the ADF in the workbook
        self.saveHelper(wb, filename)

    def saveHelper(self,wb,filename):
        """
        helper class for saveNew which provides the ability to save the
        ADF in an excel file
        
        Parameters
        ----------
        adf : class
            the instance of the ADF() class
        name : str
            the filename
        wb : xl.Workbook()
            the excel workbook to save the adf in
        """
        
        #checks to see if there is a file with the same name to overwrite or not
        try: 
            wb['{}'.format(filename)]
            ws = wb['{}'.format(filename)]
            
        except:
            wb.create_sheet('{}'.format(filename))
            ws = wb['{}'.format(filename)]
        
        for i in self.nodes.values():
            
            #duplicate flag
            dupeFlag = True
            
            #checks whether the node has already been saved or not to prevent duplicates
            for row in ws.iter_rows(values_only = True):   
                if row[0] == i.name:
                    dupeFlag = False             
                else:
                    pass
                
            if dupeFlag == True:
                
                #sets node name
                nodeName = [i.name]
                
                try:
                    #adds acceptance conditions and corresponding statements
                    for x,y in zip(i.acceptanceOriginal,i.statement):
                        nodeName.append(x)
                        nodeName.append(y)
                        
                    #adds the rejection statement
                    nodeName.append(i.statement[-1])
                    
                except:
                    #no acceptance condition i.e. base level factor
                    pass
                
                ws.append(nodeName)
                
                #adds the question in the row below the other node info
                if i.question != None:
                    ws.append([i.question])
                    
                else:
                    ws.append(['None'])
        
        #inserts a row at the end with the question order
        if self.questionOrder != []:   
            self.questionOrder.insert(0,"QUESTION!")
            ws.append(self.questionOrder)
        
        wb.save('{}.xlsx'.format(filename))  
        
        #removes from the question order as only temporary for data file
        self.questionOrder.remove("QUESTION!")
        
class Node:
    """
    A class used to represent an individual node, whose acceptance conditions
    are instantiated by 'yes' or 'no' questions

    Attributes
    ----------
    name : str
        the name of the node
    question : str, optional
        the question which will instantiate the blf
    answers : 
        set to None type to indicate to other methods the Node is not from MultiChoice()
    acceptanceOriginal : str
        the original acceptance condition before being converted to postfix notation
    statement : list
        the statements which will be output depending on whether the node is accepted or rejected
    acceptance : list
        the acceptance condition in postfix form
    children : list
        a list of the node's children nodes
    
    Methods
    -------
    attributes(acceptance)
        sets the acceptance conditions and determines the children nodes  
    
    logicConverter(expression)
        converts the acceptance conditions into postfix notation
        
    """
    def __init__(self, name, acceptance=None, statement=None, question=None):
        """
        Parameters
        ----------
        name : str
            the name of the node
        statement : list, optional
            the statements which will be output depending on whether the node is accepted or rejected
        acceptance : list, optional
            the acceptance condition in postfix form
        question : str, optional
            the question which will instantiate the blf
        """
        #name of the node
        self.name = name
        
        #question for base leve factor
        self.question = question
        
        self.answers = None
        
        self.acceptanceOriginal = acceptance
    
        #sets postfix acceptance conditions and children nodes
        try:
            self.attributes(acceptance)
            self.statement = statement
        except:
            self.acceptance = None
            self.children = None
            self.statement = None
    
    def attributes(self, acceptance):
        """
        sets the acceptance condition and children for the node
        
        Parameters
        ----------
        acceptance : list
            the acceptance condition in postfix form
        """
        
        #sets acceptance condition to postfix if acceptance condition specified
        self.acceptance = []
        self.children = []
        
        for i in acceptance:
            self.acceptance.append(self.logicConverter(i))
        
        for i in self.acceptance:
            splitAcceptance = i.split()
            
            #sets the children nodes
            for token in splitAcceptance:
                
                if token not in ['and','or','not','reject'] and token not in self.children:
                    
                    self.children.append(token)   

    def logicConverter(self, expression):
        """
        converts a logical expression from infix to postfix notation
        
        Parameters
        ----------
        expression : list
            the acceptance condition to be converted into postfix form
        """
        
        #precedent dictionary of logical operators and reject keyword
        precedent = {'(':1,'or':2,'and':3,'not':4,'reject':5}
        
        #creates the stack
        operatorStack = Stack()
        
        #splits the tokens in the logical expression
        tokenList = expression.split()
        
        #stores the postfix expression
        postfixList = []

        #checks each token in the expression and pushes or pops on the stack accordingly
        for token in tokenList:
            
            if token == '(':
                operatorStack.push(token)
            elif token == ')':
                topToken = operatorStack.pop()
                while topToken != '(':
                    postfixList.append(topToken)
                    topToken = operatorStack.pop()

                                
            elif token == 'and' or token == 'or' or token == 'not' or token == 'reject':
                while (not operatorStack.isEmpty()) and (precedent[operatorStack.peek()] >= precedent[token]):
                    postfixList.append(operatorStack.pop())
                operatorStack.push(token)
                
            else:
                postfixList.append(token)

        #while operator stack not empty pop the operators to the postfix list
        while not operatorStack.isEmpty():
            postfixList.append(operatorStack.pop())
        
        #returns the post fix expression as a string  
        return " ".join(postfixList)

class MultiChoice(Node):
    """
    for the creation of multiple choice base level factors, especially to 
    facilitate the exception to the 4th amendment and NIHL domains
    
    Methods inherited from Node()
    
    Attributes
    ----------
    name : str
        the name of the node
    question : str, optional
        the question which will instantiate the blf
    answers : list
        the multiple choice answers to be selected from
    acceptanceOriginal : str
        the original acceptance condition before being converted to postfix notation
    statement : list
        the statements which will be output depending on whether the node is accepted or rejected
    acceptance : list
        the acceptance condition in postfix form
    children : list
        a list of the node's children nodes
        
    Methods
    -------
    attributes(acceptance)
        sets the acceptance conditions and determines the children nodes  
    
    logicConverter(expression)
        converts the acceptance conditions into postfix notation
    
    """
    def __init__(self, name, acceptance, statement, question=None):
        """
        Parameters
        ----------
        name : str
            the name of the node
        statement : list, optional
            the statements which will be output depending on whether the node is accepted or rejected
        acceptance : list, optional
            the acceptance condition in postfix form
        question : str, optional
            the question which will instantiate the blf
        """
        
        #name of the node
        self.name = name
        
        #quetion for base level factor
        self.question = question
        
        self.acceptanceOriginal = acceptance

        #sets postfix acceptance conditions and children nodes
        try:
            self.attributes(acceptance)
            self.statement = statement
        except:
            self.acceptance = None
            self.children = None
            self.statement = None
                
        self.answers = self.children
 
def importADF(file,name):
    """
    enables a .xlsx file containing the ADF to be loaded into the tool
    
    Parameters
    ----------
    file : file
        the file address on the computer
    name : str
        the filename
    """
    
    try:

        #loads workbook
        wb = load_workbook(file)
        
        #creates ADF
        ws = wb[name]
        adf = ADF(name)
        
        #identifies question order row and converts to questionOrder attribute
        if ws[ws.max_row][0].value == 'QUESTION!':
            
            max_range = ws.max_row - 1
            adf.questionOrder = [x.value for x in ws[ws.max_row] if x.value is not None]
            adf.questionOrder.remove("QUESTION!")
           
        else:
            max_range = ws.max_row
        
        for row in range(1, max_range,2):
            
            #each node has two rows with its data
            firstrow = ws[row]
            secondrow = ws[row+1]
            
            row = [x.value for x in firstrow if x.value is not None]
            row1 = [x.value for x in secondrow if x.value is not None]
            
            nameNode = row[0]
            acceptance = row[1:-1:2]
            slice = (row[0::2])
            statement = slice[1:]
            if statement != []:
                statement.append(row[-1])
            question = row1[0]

            #checks whether a multiChoice() or Node() is more appropriate
            if row1[0] == 'None':
                adf.addNodes(nameNode,acceptance,statement)
                
            else:
                
                try:
                    #if row[1] is present this indicates there is an acceptance condition so it is a multiple choice node since there is a question
                    row[1]
                    adf.addMulti(nameNode,acceptance,statement,question)   
                     
                except:
                    adf.addNodes(nameNode,acceptance,statement,question)         
               
        return adf
    
    except IOError:
        print('\nFilename does not exist\n')
